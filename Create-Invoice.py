# library to import the excel file
import openpyxl
from openpyxl import load_workbook

# libraries to create the pdf file and add text to it
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl import load_workbook
# library to get logo related information
from PIL import Image

# convert the font so it is compatible
pdfmetrics.registerFont(TTFont('Arial', 'Arial.ttf'))

# import the sheet from the excel file
wb = openpyxl.load_workbook('C://Users//LENOVO//PycharmProjects//Creating_Invoices//Invoice_Record.xlsx')
sheet = wb['Invoices']

# import company's logo
im = Image.open('logo.jpg')
width, height = im.size
ratio = width / height
image_width = 800
image_height = int(image_width / ratio)

# Page information
page_width = 2156
page_height = 3050

# Invoice variables
company_name = 'The best company in the world'
payment_terms = 'x'
contact_info = 'x'
margin = 100
month_year = 'August 2020'


# def function
def create_invoice():
    for i in range(2, 12):
        # Reading values from excel file

        invoice_number = sheet.cell(row=i, column=1).value
        Client_name = sheet.cell(row=i, column=2).value
        Client_email = sheet.cell(row=i, column=3).value
        Project_name = sheet.cell(row=i, column=4).value.lower()


        # Creating a pdf file and setting a naming convention
        c = canvas.Canvas(str(invoice_number) + '_' + Project_name + '.pdf')

        c.setPageSize((page_width, page_height))




        # Drawing the image
        c.drawInlineImage("C:\\Users\\LENOVO\\PycharmProjects\\Creating_Invoices\\logo.jpg", page_width - image_width - margin,
                          page_height - image_height - margin,
                          image_width, image_height)

        # Invoice information
        c.setFont('Arial', 80)
        text = 'INVOICE'
        text_width = stringWidth(text, 'Arial', 80)
        c.drawString((page_width - text_width) / 2, page_height - image_height - margin, text)
        y = page_height - image_height - margin * 4
        x = 2 * margin
        x2 = x + 550

        c.setFont('Arial', 45)
        c.drawString(x, y, 'invoice number: ')
        c.drawString(x2, y, invoice_number)
        y -= margin

        c.drawString(x, y, 'Issued to: ')
        c.drawString(x2, y, Client_name)
        y -= margin

        c.drawString(x, y, 'Client email id: ')
        c.drawString(x2, y, str(Client_email))
        y -= margin

        c.drawString(x, y, 'Project name: ')
        c.drawString(x2, y, Project_name)
        y -= margin


        # Saving the pdf file
        c.save()

create_invoice()

